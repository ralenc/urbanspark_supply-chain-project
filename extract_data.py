import json

from openpyxl import load_workbook


def extract_data(workbook_path):
    """
    Extract per-SKU inventory policy parameters and 12-month demand history
    from the UrbanSpark Excel workbook.

    Parameters:
        workbook_path: path to UrbanSpark_RawData_AsIs.xlsx

    Returns:
        inv_opt_data.json  — policy parameters (main table + second table)
        demand_history.json — 12-month demand per SKU from Order_History_2024
    """
    wb = load_workbook(workbook_path, data_only=True)

    # ── Inventory_Optimization — main table (rows 2-39) ───────────────────────
    ws_inv = wb["Inventory_Optimization"]
    main = {}
    for r in range(2, 40):
        sku = ws_inv.cell(r, 1).value
        if not sku:
            continue
        main[sku] = {
            "ABC_XYZ": ws_inv.cell(r, 3).value,
            "LT": ws_inv.cell(r, 4).value,
            "AvgMonthly": ws_inv.cell(r, 5).value,
            "StdDev": ws_inv.cell(r, 6).value,
            "Z": ws_inv.cell(r, 7).value,
            "NewSS": ws_inv.cell(r, 8).value,
            "NewROP": ws_inv.cell(r, 9).value,
            "AnnualDemand": ws_inv.cell(r, 10).value,
            "PurchCost": ws_inv.cell(r, 11).value,
            "EOQ": ws_inv.cell(r, 12).value,
        }

    # ── Inventory_Optimization — second table (rows 42-79) ────────────────────
    second = {}
    for r in range(42, 80):
        sku = ws_inv.cell(r, 1).value
        if not sku:
            continue
        second[sku] = {
            "OldSS": ws_inv.cell(r, 2).value,
            "OldROP": ws_inv.cell(r, 4).value,
            "OldOrderQty": ws_inv.cell(r, 6).value,
        }

    # ── Order_History_2024 — monthly demand columns ───────────────────────────
    ws_oh = wb["Order_History_2024"]

    # Find month columns (headers containing '-24')
    headers = [ws_oh.cell(1, c).value for c in range(1, 20)]
    month_cols = [
        c
        for c, h in enumerate(headers, start=1)
        if h and "-24" in str(h) and h != "Total_2024"
    ]

    demand = {}
    for r in range(2, 40):
        sku = ws_oh.cell(r, 1).value
        if not sku:
            continue
        demand[sku] = [ws_oh.cell(r, c).value for c in month_cols]

    # ── Save to JSON ──────────────────────────────────────────────────────────
    with open("inv_opt_data.json", "w") as f:
        json.dump({"main": main, "second": second}, f, indent=2, default=str)

    with open("demand_history.json", "w") as f:
        json.dump(demand, f, indent=2)

    print(f"Extracted {len(main)} SKUs from Inventory_Optimization")
    print(f"Extracted {len(demand)} SKUs from Order_History_2024")
    print("Saved inv_opt_data.json and demand_history.json")


if __name__ == "__main__":
    extract_data("UrbanSpark_RawData_AsIs.xlsx")
